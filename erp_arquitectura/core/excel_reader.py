"""
core/services/excel_reader.py
══════════════════════════════════════════════════════════════════════════════
Lector de Excel para datos del Formulario 12.4 MINVU.
Incluye generador de plantilla Excel para entregar a los arquitectos.
══════════════════════════════════════════════════════════════════════════════
"""

from decimal import Decimal, InvalidOperation
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .minvu_pdf_filler import DatosFormulario124


# ══════════════════════════════════════════════════════════════════════════════
# MAPEO EXCEL → DatosFormulario124
# Formato: "NombreHoja": { "CeldaExcel": "campo_en_DatosFormulario124" }
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_MAP = {
    "1_Propiedad": {
        "B3": "municipalidad",
        "B4": "region",
        "B5": "calle_inmueble",
        "B6": "numero_inmueble",
        "B7": "rol_sii",
        "B8": "rol_avaluo",
        "B9": "lote",
        "B10": "loteo_localidad",
        "B11": "conservador_bienes_raices",
        "B12": "inscrito_fojas",
        "B13": "inscrito_numero",
        "B14": "anio_inscripcion",
        "B15": "registro_propiedad_anio",
    },
    "2_Propietario": {
        "B3": "propietario_nombre",
        "B4": "propietario_rut",
        "B5": "propietario_empresa",
        "B6": "propietario_empresa_rut",
        "B7": "propietario_rep_legal",
        "B8": "propietario_calle",
        "B9": "propietario_num",
        "B10": "propietario_comuna",
        "B11": "propietario_email",
        "B12": "propietario_telefono",
        "B13": "propietario_celular",
        # También como declarante DJ
        "B16": "declarante_nombre",
        "B17": "declarante_ci",
        "B18": "declarante_calle",
        "B19": "declarante_numero",
    },
    "3_Arquitecto": {
        "B3": "arq_empresa_nombre",
        "B4": "arq_empresa_rut",
        "B5": "arq_nombre",
        "B6": "arq_rut",
        "B7": "arq_profesion",
        "B8": "arq_patente",
        "B9": "arq_calle",
        "B10": "arq_numero",
        "B11": "arq_comuna",
    },
    "4_Superficies": {
        "B3": "_sup_existente_str",       # campo especial → se convierte a Decimal
        "B4": "_sup_regularizar_str",     # campo especial → se convierte a Decimal
        "B5": "superficie_terreno_m2",
        "B6": "sup_1p_con_permiso",
        "B7": "sup_1p_regularizar",
        "B8": "sup_2p_con_permiso",
        "B9": "sup_2p_regularizar",
        "B10": "sup_3p_con_permiso",
        "B11": "sup_3p_regularizar",
    },
    "5_Avaluo": {
        "B3": "avaluo_uf",
        "B4": "avaluo_pesos",
        "B5": "avaluo_terreno_uf",
        "B6": "tipo_agrupamiento",
        "B7": "altura_max_permitida",
        "B8": "clasificacion_predominante",
    },
    "6_Permisos": {
        "B3": "_tiene_permiso_ant",       # campo especial: SI/NO
        "B4": "permiso_anterior_num",
        "B5": "permiso_anterior_anio",
        "B6": "_tiene_recepcion_ant",     # campo especial: SI/NO
        "B7": "recepcion_anterior_num",
        "B8": "recepcion_anterior_anio",
        "B9": "_en_copropiedad",          # campo especial: SI/NO
    },
    "7_Normas": {
        "B3": "sist_agrup_permitido",
        "B4": "sist_agrup_existente",
        "B5": "distanciamiento_permitido",
        "B6": "distanciamiento_existente",
        "B7": "rasante_existente",
        "B8": "altura_cierros_existente",
        "B9": "densidad_existente",
        "B10": "adosamiento_norma",
        "B11": "forma_cumplimiento_art70",
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# LECTOR DE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

class ExcelReader124:
    """Lee el Excel de datos y retorna un DatosFormulario124."""

    def leer(self, excel_path: Path) -> tuple[DatosFormulario124, list[str]]:
        """
        Returns:
            (datos, errores) — errores es lista de strings describiendo problemas.
        """
        errores = []
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        raw = {}

        for hoja_nombre, campos in EXCEL_MAP.items():
            if hoja_nombre not in wb.sheetnames:
                errores.append(f"Hoja '{hoja_nombre}' no encontrada en el Excel.")
                continue
            ws = wb[hoja_nombre]
            for celda, campo in campos.items():
                valor = ws[celda].value
                raw[campo] = "" if valor is None else str(valor).strip()

        # Construir DatosFormulario124
        datos = DatosFormulario124()

        # Campos directos
        for campo, valor in raw.items():
            if not campo.startswith("_") and hasattr(datos, campo):
                setattr(datos, campo, valor)

        # Campos especiales: superficies
        try:
            sup_e = raw.get("_sup_existente_str", "0").replace(",", ".")
            datos.sup_existente_decimal = Decimal(sup_e) if sup_e else Decimal("0")
        except InvalidOperation:
            errores.append("Superficie existente no es un número válido.")

        try:
            sup_r = raw.get("_sup_regularizar_str", "0").replace(",", ".")
            datos.sup_regularizar_decimal = Decimal(sup_r) if sup_r else Decimal("0")
        except InvalidOperation:
            errores.append("Superficie a regularizar no es un número válido.")

        # Campos booleanos (SI/NO en Excel)
        def es_si(val: str) -> bool:
            return val.upper().strip() in ("SI", "SÍ", "S", "YES", "1", "TRUE")

        datos.tiene_permiso_anterior   = es_si(raw.get("_tiene_permiso_ant", "NO"))
        datos.tiene_recepcion_anterior = es_si(raw.get("_tiene_recepcion_ant", "NO"))
        datos.en_copropiedad           = es_si(raw.get("_en_copropiedad", "NO"))

        return datos, errores


# ══════════════════════════════════════════════════════════════════════════════
# GENERADOR DE PLANTILLA EXCEL
# ══════════════════════════════════════════════════════════════════════════════

class ExcelTemplateGenerator:
    """
    Genera el Excel plantilla para que el arquitecto ingrese los datos.
    Listo para distribuir al equipo.
    """

    COLOR_HEADER    = "1F3864"  # azul oscuro
    COLOR_SECCION   = "2F75B6"  # azul medio
    COLOR_INPUT     = "FFFFFF"  # blanco (campos a llenar)
    COLOR_CALCULADO = "E2EFDA"  # verde claro (calculado automáticamente)
    COLOR_INFO      = "FFF2CC"  # amarillo claro (informativo)

    def generar(self, output_path: Path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remover hoja vacía default

        self._hoja_instrucciones(wb)
        self._hoja_propiedad(wb)
        self._hoja_propietario(wb)
        self._hoja_arquitecto(wb)
        self._hoja_superficies(wb)
        self._hoja_avaluo(wb)
        self._hoja_permisos(wb)
        self._hoja_normas(wb)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

    # ── Estilos helpers ───────────────────────────────────────────────────────

    def _header_style(self, ws, row, texto, color=None):
        color = color or self.COLOR_SECCION
        cell = ws.cell(row=row, column=1, value=texto)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor=color)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell.alignment = Alignment(horizontal="center")

    def _campo(self, ws, row, label, ejemplo="", requerido=True, calculado=False):
        color_label = "F2F2F2"
        color_val   = self.COLOR_CALCULADO if calculado else self.COLOR_INPUT
        req = " *" if requerido else ""

        c_label = ws.cell(row=row, column=1, value=f"{label}{req}")
        c_label.font = Font(bold=True, size=10)
        c_label.fill = PatternFill("solid", fgColor=color_label)
        c_label.alignment = Alignment(wrap_text=True)

        c_val = ws.cell(row=row, column=2, value=ejemplo)
        c_val.fill = PatternFill("solid", fgColor=color_val)
        c_val.alignment = Alignment(horizontal="left", vertical="center")
        c_val.border = Border(
            bottom=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        )

        c_nota = ws.cell(row=row, column=3, value="← INGRESE AQUÍ" if not calculado else "Auto-calculado")
        c_nota.font = Font(italic=True, color="808080", size=9)

        ws.row_dimensions[row].height = 20

    def _setup_columnas(self, ws):
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 20

    def _hoja_instrucciones(self, wb):
        ws = wb.create_sheet("📋 INSTRUCCIONES")
        ws.column_dimensions["A"].width = 80
        instrucciones = [
            ("FORMULARIO 12.4 — LEY 20.898 TÍTULO I", True, self.COLOR_HEADER),
            ("PLANTILLA DE DATOS PARA LLENADO AUTOMÁTICO DE PDF", True, self.COLOR_SECCION),
            ("", False, None),
            ("INSTRUCCIONES DE USO:", True, None),
            ("1. Complete TODAS las hojas marcadas con * (obligatorio).", False, None),
            ("2. Los campos con fondo VERDE se calculan automáticamente, no los edite.", False, None),
            ("3. Para campos SI/NO, escriba exactamente: SI o NO", False, None),
            ("4. Superficies con coma decimal: ej. 72,37", False, None),
            ("5. RUT formato: 12.345.678-9", False, None),
            ("6. Avalúo fiscal: usar valor a la fecha 04/02/2016 (publicación Ley 20.898)", False, None),
            ("", False, None),
            ("HOJAS DEL ARCHIVO:", True, None),
            ("  1_Propiedad      → Datos del predio (rol, dirección, municipalidad)", False, None),
            ("  2_Propietario    → Datos del dueño del inmueble", False, None),
            ("  3_Arquitecto     → Datos del profesional que suscribe", False, None),
            ("  4_Superficies    → m² existentes, a regularizar, terreno", False, None),
            ("  5_Avaluo         → Avalúo fiscal y tipo de agrupamiento", False, None),
            ("  6_Permisos       → Permisos anteriores (si los hay)", False, None),
            ("  7_Normas         → Normas urbanísticas para sección 5.7", False, None),
            ("", False, None),
            ("⚠ LÍMITES LEY 20.898 ART. 3°:", True, None),
            ("  • Superficie TOTAL no puede superar 140 m²", False, None),
            ("  • Avalúo fiscal no puede superar 2.000 UF (al 04/02/2016)", False, None),
        ]
        for i, (texto, bold, color) in enumerate(instrucciones, start=1):
            c = ws.cell(row=i, column=1, value=texto)
            c.font = Font(bold=bold, size=11 if bold else 10,
                          color="FFFFFF" if color else "000000")
            if color:
                c.fill = PatternFill("solid", fgColor=color)
                c.alignment = Alignment(horizontal="center")
            ws.row_dimensions[i].height = 22

    def _hoja_propiedad(self, wb):
        ws = wb.create_sheet("1_Propiedad")
        self._setup_columnas(ws)
        self._header_style(ws, 1, "SECCIÓN 1 — DATOS DE LA PROPIEDAD")
        ws.cell(row=2, column=1, value="* = Obligatorio").font = Font(italic=True, color="FF0000")
        campos = [
            (3,  "Municipalidad DOM *",              "CURACAVÍ",             True),
            (4,  "Región *",                         "Metropolitana",        True),
            (5,  "Calle / Camino *",                 "RUTA G-68",            True),
            (6,  "Número *",                         "A-1",                  True),
            (7,  "Rol SII *",                        "104-273",              True),
            (8,  "Rol de Avalúo *",                  "104-273",              True),
            (9,  "Lote",                             "14",                   False),
            (10, "Loteo / Localidad",                "BATALLA SAN JUAN",     False),
            (11, "Conservador de Bienes Raíces *",   "CURACAVÍ",             True),
            (12, "Inscrito a Fojas Nº",              "CURACAVÍ",             False),
            (13, "Inscrito Número",                  "1383",                 False),
            (14, "Año Inscripción",                  "1652",                 False),
            (15, "Año Registro Propiedad",           "2022",                 False),
        ]
        for row, label, ejemplo, req in campos:
            self._campo(ws, row, label, ejemplo, req)

    def _hoja_propietario(self, wb):
        ws = wb.create_sheet("2_Propietario")
        self._setup_columnas(ws)
        self._header_style(ws, 1, "SECCIÓN 3 — DATOS DEL PROPIETARIO")
        campos = [
            (3,  "Nombre Propietario *",              "CRISTIAN SILVA HERNANDEZ", True),
            (4,  "RUT Propietario *",                 "10.782.449-9",             True),
            (5,  "Empresa (si corresponde)",          "TERRAEM SPA",              False),
            (6,  "RUT Empresa",                       "77.796.999-4",             False),
            (7,  "Representante Legal",               "",                         False),
            (8,  "Calle Propietario *",               "RUTA G-68",                True),
            (9,  "Número *",                          "A-1",                      True),
            (10, "Comuna *",                          "CURACAVÍ",                 True),
            (11, "Email",                             "contacto@terraem.cl",      False),
            (12, "Teléfono",                          "",                         False),
            (13, "Celular",                           "+56998265930",             False),
        ]
        self._header_style(ws, 14, "DECLARANTE (Sección DJ — generalmente = Propietario)", self.COLOR_HEADER)
        campos += [
            (16, "Nombre Declarante *",               "CRISTIAN SILVA HERNANDEZ", True),
            (17, "Cédula de Identidad *",             "10.782.449-9",             True),
            (18, "Calle Declarante",                  "RUTA G-68",                False),
            (19, "Número Declarante",                 "A-1",                      False),
        ]
        for row, label, ejemplo, req in campos:
            self._campo(ws, row, label, ejemplo, req)

    def _hoja_arquitecto(self, wb):
        ws = wb.create_sheet("3_Arquitecto")
        self._setup_columnas(ws)
        self._header_style(ws, 1, "SECCIÓN 4 — DATOS DEL ARQUITECTO")
        campos = [
            (3,  "Nombre Empresa *",          "TERRAEM SPA",          True),
            (4,  "RUT Empresa *",             "77.796.999-4",         True),
            (5,  "Nombre Arquitecto *",       "LORETO SALAZAR SALINA",True),
            (6,  "RUT Arquitecto *",          "16.531.751-3",         True),
            (7,  "Profesión *",               "ARQUITECTO",           True),
            (8,  "Patente Nº *",              "300833",               True),
            (9,  "Calle *",                   "RUTA G-708",           True),
            (10, "Número *",                  "571",                  True),
            (11, "Comuna *",                  "MELIPILLA",            True),
        ]
        for row, label, ejemplo, req in campos:
            self._campo(ws, row, label, ejemplo, req)

    def _hoja_superficies(self, wb):
        ws = wb.create_sheet("4_Superficies")
        self._setup_columnas(ws)
        self._header_style(ws, 1, "SECCIÓN 5.3 — SUPERFICIES (m²)")
        ws.cell(row=2, column=1,
                value="⚠ LÍMITE: Superficie total NO puede superar 140 m² (Ley 20.898 Art. 3°)"
               ).font = Font(bold=True, color="FF0000")
        campos = [
            (3,  "Superficie existente TOTAL (m²) *",     "72,37",  True,  False),
            (4,  "Superficie a regularizar (m²) *",       "33,67",  True,  False),
            (5,  "Superficie del terreno (m²) *",         "28,03",  True,  False),
            (6,  "1er Piso — con permiso (m²)",           "72,37",  False, False),
            (7,  "1er Piso — a regularizar (m²)",         "--",     False, False),
            (8,  "2do Piso — con permiso (m²)",           "--",     False, False),
            (9,  "2do Piso — a regularizar (m²)",         "15,11",  False, False),
            (10, "3er Piso — con permiso (m²)",           "",       False, False),
            (11, "3er Piso — a regularizar (m²)",         "",       False, False),
        ]
        for row, label, ejemplo, req, calc in campos:
            self._campo(ws, row, label, ejemplo, req, calc)
        ws.cell(row=12, column=1, value="TOTAL calculado automáticamente →").font = Font(italic=True, color="1F7A1F")

    def _hoja_avaluo(self, wb):
        ws = wb.create_sheet("5_Avaluo")
        self._setup_columnas(ws)
        self._header_style(ws, 1, "SECCIÓN 5.5 — AVALÚO FISCAL")
        ws.cell(row=2, column=1,
                value="Avalúo fiscal VIGENTE al 04/02/2016 (fecha publicación Ley 20.898)"
               ).font = Font(italic=True, color="C00000")
        campos = [
            (3,  "Avalúo fiscal (UF) *",              "34,48",    True),
            (4,  "Avalúo fiscal ($) *",               "38.344.495",True),
            (5,  "Avalúo terreno (UF, fecha solicitud)","965,31",  False),
            (6,  "Tipo agrupamiento *",               "AISLADO",  True),
            (7,  "Altura máxima permitida",           "2 PISOS",  False),
            (8,  "Clasificación predominante",        "E",        False),
        ]
        for row, label, ejemplo, req in campos:
            self._campo(ws, row, label, ejemplo, req)
        ws.cell(row=10, column=1, value="Tipo agrupamiento: AISLADO | PAREADO | CONTINUO | ADOSADO"
               ).font = Font(italic=True, color="808080")

    def _hoja_permisos(self, wb):
        ws = wb.create_sheet("6_Permisos")
        self._setup_columnas(ws)
        self._header_style(ws, 1, "SECCIÓN 5.2 — PERMISOS ANTERIORES")
        campos = [
            (3,  "¿Tiene permiso anterior? (SI/NO) *",    "NO",  True),
            (4,  "Permiso anterior Nº",                   "",    False),
            (5,  "Año permiso anterior",                  "",    False),
            (6,  "¿Tiene recepción anterior? (SI/NO) *",  "NO",  True),
            (7,  "Recepción anterior Nº",                 "",    False),
            (8,  "Año recepción anterior",                "",    False),
            (9,  "¿En copropiedad? (SI/NO) *",            "NO",  True),
        ]
        for row, label, ejemplo, req in campos:
            self._campo(ws, row, label, ejemplo, req)

    def _hoja_normas(self, wb):
        ws = wb.create_sheet("7_Normas")
        self._setup_columnas(ws)
        self._header_style(ws, 1, "SECCIÓN 5.7 — NORMAS URBANÍSTICAS")
        ws.cell(row=2, column=1,
                value="Completar si se conocen. El sistema calcula densidad y cesión."
               ).font = Font(italic=True, color="808080")
        campos = [
            (3,  "Sistema agrupamiento permitido",  "--",     False),
            (4,  "Sistema agrupamiento existente",  "0.05",   False),
            (5,  "Distanciamiento permitido",       "--",     False),
            (6,  "Distanciamiento existente",       "0,04",   False),
            (7,  "Rasante existente",               "40%",    False),
            (8,  "Altura cierros existente",        "--",     False),
            (9,  "Densidad existente",              "100 H/H",False),
            (10, "Adosamientos norma",              "Art. 2.6.3", False),
            (11, "Forma cumplimiento Art. 70 *",    "CESION", True),
        ]
        for row, label, ejemplo, req in campos:
            self._campo(ws, row, label, ejemplo, req)
        ws.cell(row=12, column=1, value="Forma Art. 70: CESION | APORTE | OTRO"
               ).font = Font(italic=True, color="808080")
