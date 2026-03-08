from __future__ import annotations
"""
core/services/excel_reader.py
Lector de Excel para Formulario 12.4 MINVU — Ley 20.898 Art. 3° — hasta 140 m²
"""
from decimal import Decimal, InvalidOperation
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .minvu_pdf_filler import DatosFormulario124

EXCEL_MAP = {
    "1_Propiedad": {
        "B3": "municipalidad",
        "B4": "region",
        "B5": "calle_inmueble",
        "B6": "numero_inmueble",
        "B7": "rol_sii",
        "B8": "rol_avaluo",
        "B9": "lote_excel",
        "B10": "loteo_localidad",
        "B11": "manzana",
        "B12": "conservador_bienes_raices",
        "B13": "inscrito_fojas",
        "B14": "inscrito_numero",
        "B15": "anio_inscripcion",
        "B16": "registro_propiedad_anio",
    },
    "2_Propietario": {
        "B3": "propietario_nombre",
        "B4": "propietario_rut",
        "B5": "propietario_empresa",
        "B6": "propietario_empresa_rut",
        "B7": "propietario_rep_legal",
        "B8": "propietario_rep_rut",
        "B9": "propietario_calle",
        "B10": "propietario_num",
        "B11": "propietario_comuna",
        "B12": "propietario_email",
        "B13": "propietario_telefono",
        "B14": "propietario_celular",
        "B16": "declarante_nombre",
        "B17": "declarante_ci",
        "B18": "declarante_calle",
        "B19": "declarante_numero",
    },
    "3_Arquitecto": {
        "B3": "arq_empresa_nombre",
        "B4": "arq_empresa_rut",
        "B5": "arq_nombre",
        "B6": "arq_rut",
        "B7": "arq_profesion",
        "B8": "arq_patente",
        "B9": "arq_calle",
        "B10": "arq_numero",
        "B11": "arq_comuna",
    },
    "4_Superficies": {
        "B3": "_sup_1p_con_permiso",
        "B4": "_sup_1p_regularizar",
        "B5": "_sup_2p_con_permiso",
        "B6": "_sup_2p_regularizar",
        "B7": "_sup_3p_con_permiso",
        "B8": "_sup_3p_regularizar",
        "B9": "superficie_terreno_m2",
    },
    "5_Avaluo": {
        "B3": "avaluo_uf",
        "B4": "avaluo_pesos",
        "B5": "avaluo_terreno_uf",
        "B6": "tipo_agrupamiento",
        "B7": "altura_max_permitida",
        "B8": "clasificacion_1",
        "B9": "clasificacion_1_m2",
    },
    "6_Permisos": {
        "B3": "_tiene_permiso_ant",
        "B4": "permiso_anterior_num",
        "B5": "permiso_anterior_anio",
        "B6": "_tiene_recepcion_ant",
        "B7": "recepcion_anterior_num",
        "B8": "recepcion_anterior_anio",
        "B9": "_en_copropiedad",
    },
    "7_Normas": {
        "B3": "sist_agrup_permitido",
        "B4": "sist_agrup_existente",
        "B5": "distanciamiento_permitido",
        "B6": "distanciamiento_existente",
        "B7": "rasante_existente",
        "B8": "altura_cierros_existente",
        "B9": "densidad_existente",
        "B10": "coef_constructibilidad_permitido",
        "B11": "coef_constructibilidad_existente",
        "B12": "forma_cumplimiento_art70",
    },
}


class ExcelReader124:
    def leer(self, excel_path: Path) -> tuple[DatosFormulario124, list[str]]:
        errores = []
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        raw = {}

        for hoja_nombre, campos in EXCEL_MAP.items():
            if hoja_nombre not in wb.sheetnames:
                errores.append(f"Hoja '{hoja_nombre}' no encontrada en el Excel.")
                continue
            ws = wb[hoja_nombre]
            for celda, campo in campos.items():
                valor = ws[celda].value
                raw[campo] = "" if valor is None else str(valor).strip()

        datos = DatosFormulario124()

        for campo, valor in raw.items():
            if not campo.startswith("_") and hasattr(datos, campo):
                setattr(datos, campo, valor)

        # Superficies con cálculo de totales
        def to_dec(val):
            try:
                return Decimal(val.replace(",", ".")) if val and val not in ("--", "") else Decimal("0")
            except:
                return Decimal("0")

        s1cp = to_dec(raw.get("_sup_1p_con_permiso", ""))
        s1rg = to_dec(raw.get("_sup_1p_regularizar", ""))
        s2cp = to_dec(raw.get("_sup_2p_con_permiso", ""))
        s2rg = to_dec(raw.get("_sup_2p_regularizar", ""))
        s3cp = to_dec(raw.get("_sup_3p_con_permiso", ""))
        s3rg = to_dec(raw.get("_sup_3p_regularizar", ""))

        def fmt(d): return str(d).rstrip('0').rstrip('.') if d else ""

        datos.sup_1p_con_permiso = fmt(s1cp)
        datos.sup_1p_regularizar = fmt(s1rg)
        datos.sup_1p_total = fmt(s1cp + s1rg)
        datos.sup_2p_con_permiso = fmt(s2cp)
        datos.sup_2p_regularizar = fmt(s2rg)
        datos.sup_2p_total = fmt(s2cp + s2rg)
        datos.sup_3p_con_permiso = fmt(s3cp)
        datos.sup_3p_regularizar = fmt(s3rg)
        datos.sup_3p_total = fmt(s3cp + s3rg)
        total_cp = s1cp + s2cp + s3cp
        total_rg = s1rg + s2rg + s3rg
        total_tot = total_cp + total_rg
        datos.sup_total_con_permiso = fmt(total_cp)
        datos.sup_total_regularizar = fmt(total_rg)
        datos.sup_total_total = fmt(total_tot)

        # Validar límite 140 m²
        if total_tot > Decimal("140"):
            errores.append(f"⚠ Superficie total ({total_tot} m²) supera el límite de 140 m² de la Ley 20.898 Art. 3°.")

        # Booleanos
        def es_si(val):
            return str(val).upper().strip() in ("SI", "SÍ", "S", "YES", "1", "TRUE")

        datos.tiene_permiso_anterior   = es_si(raw.get("_tiene_permiso_ant", "NO"))
        datos.tiene_recepcion_anterior = es_si(raw.get("_tiene_recepcion_ant", "NO"))
        datos.en_copropiedad           = es_si(raw.get("_en_copropiedad", "NO"))

        # Mapear declarante = propietario si no se llenó
        if not datos.declarante_nombre:
            datos.declarante_nombre = datos.propietario_nombre
        if not datos.declarante_ci:
            datos.declarante_ci = datos.propietario_rut

        # Comprobante = mismo encabezado
        datos.comprobante_municipalidad = datos.municipalidad
        datos.comprobante_calle = datos.calle_inmueble
        datos.comprobante_numero = datos.numero_inmueble

        return datos, errores


# Alias
ExcelReader123 = ExcelReader124


class ExcelTemplateGenerator:
    COLOR_HEADER  = "1F3864"
    COLOR_SECCION = "2F75B6"
    COLOR_INPUT   = "FFFFFF"

    def generar(self, output_path: Path) -> Path:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._instrucciones(wb)
        self._propiedad(wb)
        self._propietario(wb)
        self._arquitecto(wb)
        self._superficies(wb)
        self._avaluo(wb)
        self._permisos(wb)
        self._normas(wb)
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _setup(self, ws):
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 42

    def _header(self, ws, row, texto, color=None):
        color = color or self.COLOR_HEADER
        c = ws.cell(row=row, column=1, value=texto)
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 26

    def _campo(self, ws, row, label, valor="", req=False):
        ws.cell(row=row, column=1, value=("* " if req else "  ") + label).font = Font(size=10, bold=req)
        c = ws.cell(row=row, column=2, value=valor)
        c.font = Font(size=10, color="1F3864", bold=True)
        c.fill = PatternFill("solid", fgColor="EBF3FB")
        ws.row_dimensions[row].height = 20

    def _instrucciones(self, wb):
        ws = wb.create_sheet("0_Instrucciones")
        ws.column_dimensions["A"].width = 80
        rows = [
            ("FORMULARIO 12.4 — SOLICITUD DE REGULARIZACIÓN MINVU", True, self.COLOR_HEADER),
            ("Ley 20.898 Art. 3° — Viviendas hasta 140 m² | Avalúo hasta 2.000 UF", True, self.COLOR_SECCION),
            ("", False, None),
            ("Complete cada hoja y suba este archivo al sistema ERP.", False, None),
            ("Los campos con * son obligatorios.", False, None),
            ("No modifique los nombres de las hojas.", False, None),
        ]
        for i, (t, b, c) in enumerate(rows, 1):
            cell = ws.cell(row=i, column=1, value=t)
            cell.font = Font(bold=b, size=11 if b else 10, color="FFFFFF" if c else "000000")
            if c:
                cell.fill = PatternFill("solid", fgColor=c)
            ws.row_dimensions[i].height = 22

    def _propiedad(self, wb):
        ws = wb.create_sheet("1_Propiedad")
        self._setup(ws)
        self._header(ws, 1, "1 — DATOS DE LA PROPIEDAD")
        campos = [
            (3,  "Municipalidad DOM *",            "CURACAVÍ",         True),
            (4,  "Región *",                       "Metropolitana",    True),
            (5,  "Calle / Camino *",               "RUTA G-68",        True),
            (6,  "Número *",                       "A-1",              True),
            (7,  "Rol SII *",                      "104-273",          True),
            (8,  "Rol de Avalúo *",                "104-273",          True),
            (9,  "Lote",                           "14",               False),
            (10, "Loteo / Localidad",              "BATALLA SAN JUAN", False),
            (11, "Manzana",                        "",                 False),
            (12, "Conservador de Bienes Raíces *", "CURACAVÍ",         True),
            (13, "Inscrito a Fojas Nº",            "",                 False),
            (14, "Inscrito Número",                "1383",             False),
            (15, "Año Inscripción",                "1652",             False),
            (16, "Año Registro Propiedad",         "2022",             False),
        ]
        for row, label, val, req in campos:
            self._campo(ws, row, label, val, req)

    def _propietario(self, wb):
        ws = wb.create_sheet("2_Propietario")
        self._setup(ws)
        self._header(ws, 1, "3 — DATOS DEL PROPIETARIO")
        campos = [
            (3,  "Nombre / Razón Social *",       "CRISTIAN SILVA HERNANDEZ", True),
            (4,  "RUT *",                          "10.782.449-9",             True),
            (5,  "Empresa (si corresponde)",       "TERRAEM SPA",              False),
            (6,  "RUT Empresa",                    "77.796.999-4",             False),
            (7,  "Representante Legal",            "",                         False),
            (8,  "RUT Rep. Legal",                 "",                         False),
            (9,  "Calle *",                        "RUTA G-68",                True),
            (10, "Número *",                       "A-1",                      True),
            (11, "Comuna *",                       "CURACAVÍ",                 True),
            (12, "Email",                          "contacto@terraem.cl",      False),
            (13, "Teléfono",                       "",                         False),
            (14, "Celular",                        "+56998265930",             False),
        ]
        self._header(ws, 15, "DECLARANTE DJ (generalmente = Propietario)", self.COLOR_SECCION)
        campos += [
            (16, "Nombre Declarante *",            "CRISTIAN SILVA HERNANDEZ", True),
            (17, "Cédula de Identidad *",          "10.782.449-9",             True),
            (18, "Calle Declarante",               "RUTA G-68",                False),
            (19, "Número Declarante",              "A-1",                      False),
        ]
        for row, label, val, req in campos:
            self._campo(ws, row, label, val, req)

    def _arquitecto(self, wb):
        ws = wb.create_sheet("3_Arquitecto")
        self._setup(ws)
        self._header(ws, 1, "4 — DATOS DEL ARQUITECTO / PROFESIONAL COMPETENTE")
        campos = [
            (3,  "Nombre Empresa *",          "TERRAEM SPA",           True),
            (4,  "RUT Empresa *",             "77.796.999-4",          True),
            (5,  "Nombre Arquitecto *",       "LORETO SALAZAR SALINA", True),
            (6,  "RUT *",                     "16.531.751-3",          True),
            (7,  "Profesión *",               "ARQUITECTO",            True),
            (8,  "Patente Nº *",              "300833",                True),
            (9,  "Calle *",                   "RUTA G-708",            True),
            (10, "Número *",                  "571",                   True),
            (11, "Comuna *",                  "MELIPILLA",             True),
        ]
        for row, label, val, req in campos:
            self._campo(ws, row, label, val, req)

    def _superficies(self, wb):
        ws = wb.create_sheet("4_Superficies")
        self._setup(ws)
        self._header(ws, 1, "5.3 — SUPERFICIES (m²) — LÍMITE TOTAL: 140 m²")
        ws.cell(row=2, column=1, value="Ingresa solo las superficies CON PERMISO y A REGULARIZAR. Los totales se calculan automáticamente."
               ).font = Font(italic=True, color="808080", size=9)
        campos = [
            (3,  "1er Piso — con permiso (m²)",     "72,37", False),
            (4,  "1er Piso — a regularizar (m²)",   "",      False),
            (5,  "2do Piso — con permiso (m²)",     "",      False),
            (6,  "2do Piso — a regularizar (m²)",   "33,67", False),
            (7,  "3er Piso — con permiso (m²)",     "",      False),
            (8,  "3er Piso — a regularizar (m²)",   "",      False),
            (9,  "Superficie del terreno (m²) *",   "1200",  True),
        ]
        for row, label, val, req in campos:
            self._campo(ws, row, label, val, req)

    def _avaluo(self, wb):
        ws = wb.create_sheet("5_Avaluo")
        self._setup(ws)
        self._header(ws, 1, "5.5 — AVALÚO FISCAL (vigente al 04/02/2016)")
        ws.cell(row=2, column=1, value="⚠ No puede superar 2.000 UF para acogerse a la Ley 20.898 Art. 3°"
               ).font = Font(italic=True, color="C00000", size=9)
        campos = [
            (3,  "Avalúo fiscal (UF) *",              "34,48",       True),
            (4,  "Avalúo fiscal ($) *",               "38.344.495",  True),
            (5,  "Avalúo terreno (UF, fecha solicitud)","965,31",    False),
            (6,  "Tipo agrupamiento *",               "AISLADO",     True),
            (7,  "Altura máxima permitida",           "2 PISOS",     False),
            (8,  "Clasificación construcción (ej: E)","E",           False),
            (9,  "Superficie clasificación (m²)",     "72,37",       False),
        ]
        for row, label, val, req in campos:
            self._campo(ws, row, label, val, req)
        ws.cell(row=11, column=1, value="Tipo agrupamiento: AISLADO | PAREADO | CONTINUO | ADOSAMIENTO"
               ).font = Font(italic=True, color="808080", size=9)

    def _permisos(self, wb):
        ws = wb.create_sheet("6_Permisos")
        self._setup(ws)
        self._header(ws, 1, "5.2 — PERMISOS ANTERIORES Y COPROPIEDAD")
        campos = [
            (3,  "¿Tiene permiso anterior? (SI/NO) *",    "NO", True),
            (4,  "Permiso anterior Nº",                   "",   False),
            (5,  "Año permiso anterior",                  "",   False),
            (6,  "¿Tiene recepción anterior? (SI/NO) *",  "NO", True),
            (7,  "Recepción anterior Nº",                 "",   False),
            (8,  "Año recepción anterior",                "",   False),
            (9,  "¿En copropiedad? (SI/NO) *",            "NO", True),
        ]
        for row, label, val, req in campos:
            self._campo(ws, row, label, val, req)

    def _normas(self, wb):
        ws = wb.create_sheet("7_Normas")
        self._setup(ws)
        self._header(ws, 1, "5.7 — NORMAS URBANÍSTICAS")
        campos = [
            (3,  "Sist. agrupamiento permitido",       "--",        False),
            (4,  "Sist. agrupamiento existente",       "AISLADO",   False),
            (5,  "Distanciamiento permitido",          "--",        False),
            (6,  "Distanciamiento existente",          "0,04",      False),
            (7,  "Rasante existente",                  "40%",       False),
            (8,  "Altura cierros existente",           "--",        False),
            (9,  "Densidad existente",                 "100 H/H",   False),
            (10, "Coef. constructibilidad permitido",  "--",        False),
            (11, "Coef. constructibilidad existente",  "0,05",      False),
            (12, "Forma cumplimiento Art. 70 *",       "CESION",    True),
        ]
        for row, label, val, req in campos:
            self._campo(ws, row, label, val, req)
        ws.cell(row=14, column=1, value="Art. 70: CESION | APORTE | OTRO"
               ).font = Font(italic=True, color="808080", size=9)
